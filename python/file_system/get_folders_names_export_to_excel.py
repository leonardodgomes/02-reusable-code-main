import os
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import quote_sheetname
from datetime import datetime


def get_folders_in_path(path, max_depth=None, current_depth=0):
    folders = []
    if max_depth is not None and current_depth >= max_depth:
        return folders
    
    for item in os.listdir(path):
        item_path = os.path.join(path, item)
        if os.path.isdir(item_path):
            folders.append(item)
            subfolders = get_folders_in_path(item_path, max_depth, current_depth + 1)
            folders.extend([os.path.join(item, subfolder) for subfolder in subfolders])
    return folders

def get_last_modified_date(folder_path):
    last_modified = os.path.getmtime(folder_path)
    last_modified_datetime = datetime.fromtimestamp(last_modified)
    return last_modified_datetime.strftime("%Y-%m-%d %H:%M:%S")

def export_to_excel(folders, output_file, path, max_depth=None):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Folders"

    # Set headers dynamically based on max_depth
    headers = [f"Level {i+1}" for i in range(max_depth)] if max_depth else []
    headers.append("Last Modified")
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col, value=header)

        # Apply formatting to the header cell
        font = Font(bold=True)
        cell.font = font

    # Split folder names into separate columns
    for index, folder in enumerate(folders, start=2):
        folder_parts = folder.split(os.path.sep)
        for col, folder_part in enumerate(folder_parts, start=1):
            cell = sheet.cell(row=index, column=col)
            cell.value = folder_part

            # Create a hyperlink to the folder
            cell.hyperlink = f'file://{os.path.abspath(os.path.join(path, folder))}'

    # Add last modified date in the last column
    last_modified_column = len(headers)
    for index, folder in enumerate(folders, start=2):
        last_modified_date = get_last_modified_date(os.path.join(path, folder))
        cell = sheet.cell(row=index, column=last_modified_column)
        cell.value = last_modified_date

    # Resize column widths based on content
    for column in sheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except TypeError:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column[0].column_letter].width = adjusted_width

    workbook.save(output_file)
    print("Folders exported to Excel successfully!")

# Specify the path you want to get folders from

current_directory = os.path.dirname(os.path.abspath(__file__))
current_directory = current_directory.replace('\\', '/')
print(current_directory)


name_folder = current_directory.rindex("/") + 1
root_folder = current_directory[name_folder:]
print(root_folder)


path = current_directory + '/'
print('path_folders: ' + path)

# Generate a timestamp
timestamp = datetime.now().strftime("%Y_%m_%d-%H%M")

# Specify the output file name
output_file = path + f"Folder-{root_folder}-{timestamp}.xlsx"

# Specify the maximum depth of subfolders (None for no limit)
max_depth = 4

# Get folders in the specified path
folders = get_folders_in_path(path, max_depth)

# Export folders to Excel
export_to_excel(folders, output_file, path, max_depth)
